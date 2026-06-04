from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QComboBox, QPushButton, QLabel, QDateEdit, QMessageBox, QFileDialog, QTabWidget, QTableWidget, QTableWidgetItem, QHeaderView
from PyQt5.QtCore import Qt, QDate
from database import AuditRepository, UserRepository
from views.custom_table_view import CustomTableView
from models.table_models import GenericTableModel
from i18n.translator import translate
import datetime

class AuditLogWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setLayoutDirection(Qt.RightToLeft)  # RTL للـ widget
        self.repo = AuditRepository()
        self.user_repo = UserRepository()
        layout = QVBoxLayout(self)
        
        # شريط الفلترة - RTL
        filter_layout = QHBoxLayout()
        filter_layout.setDirection(QHBoxLayout.RightToLeft)  # RTL
        
        self.user_filter = QComboBox()
        self.user_filter.addItem("الكل", None)
        for u in self.user_repo.get_all():
            self.user_filter.addItem(u['username'], u['id'])
        filter_layout.addWidget(QLabel("المستخدم:"))
        filter_layout.addWidget(self.user_filter)
        
        self.action_filter = QComboBox()
        actions = ["الكل", "إضافة قيد", "تعديل قيد", "حذف قيد", "إضافة مستخدم", "تعديل مستخدم", "حذف مستخدم", "تغيير كلمة المرور"]
        self.action_filter.addItems(actions)
        filter_layout.addWidget(QLabel("العملية:"))
        filter_layout.addWidget(self.action_filter)
        
        self.table_filter = QComboBox()
        tables = ["الكل", "expenses", "users", "settings"]
        self.table_filter.addItems(tables)
        filter_layout.addWidget(QLabel("الجدول:"))
        filter_layout.addWidget(self.table_filter)
        
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-30))
        self.start_date.setCalendarPopup(True)
        filter_layout.addWidget(QLabel("من تاريخ:"))
        filter_layout.addWidget(self.start_date)
        
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        filter_layout.addWidget(QLabel("إلى تاريخ:"))
        filter_layout.addWidget(self.end_date)
        
        apply_btn = QPushButton("تطبيق الفلتر")
        apply_btn.clicked.connect(self.refresh)
        filter_layout.addWidget(apply_btn)
        layout.addLayout(filter_layout)
        
        # أزرار إضافية - RTL
        btn_layout = QHBoxLayout()
        btn_layout.setDirection(QHBoxLayout.RightToLeft)  # RTL
        export_btn = QPushButton("📊 تصدير إلى Excel")
        export_btn.clicked.connect(self.export_to_excel)
        delete_old_btn = QPushButton("🗑 حذف السجلات القديمة (أكثر من 90 يوم)")
        delete_old_btn.clicked.connect(self.delete_old_logs)
        stats_btn = QPushButton("📈 عرض الإحصائيات")
        stats_btn.clicked.connect(self.show_stats)
        print_btn = QPushButton("🖨️ طباعة")
        print_btn.clicked.connect(self.print_audit_log)
        btn_layout.addWidget(export_btn)
        btn_layout.addWidget(delete_old_btn)
        btn_layout.addWidget(stats_btn)
        btn_layout.addWidget(print_btn)
        layout.addLayout(btn_layout)
        
        self.table = CustomTableView()
        self.table.setLayoutDirection(Qt.RightToLeft)  # RTL للجدول
        self.table.setSelectionBehavior(CustomTableView.SelectRows)
        layout.addWidget(self.table)
        
        self.refresh()
    
    def refresh(self):
        user_id = self.user_filter.currentData()
        action = self.action_filter.currentText()
        if action == "الكل":
            action = None
        table_name = self.table_filter.currentText()
        if table_name == "الكل":
            table_name = None
        start = self.start_date.date().toString("yyyy-MM-dd")
        end = self.end_date.date().toString("yyyy-MM-dd")
        logs = self.repo.get_all(limit=2000, user_id=user_id, action=action,
                                 table_name=table_name, start_date=start, end_date=end)
        data = []
        for log in logs:
            data.append({
                'id': log['id'],
                'username': log['username'],
                'action': log['action'],
                'table_name': log['table_name'],
                'record_id': log['record_id'],
                'details': log['details'],
                'ip_address': log['ip_address'] or '-',
                'timestamp': log['timestamp'][:19]
            })
        headers = ['username', 'action', 'table_name', 'record_id', 'details', 'ip_address', 'timestamp']
        display_headers = ['المستخدم', 'الإجراء', 'الجدول', 'رقم السجل', 'التفاصيل', 'عنوان IP', 'التاريخ والوقت']
        data_keys = ['username', 'action', 'table_name', 'record_id', 'details', 'ip_address', 'timestamp']
        self.model = GenericTableModel(data, display_headers, key_fields=['id'], data_keys=data_keys)
        self.table.setModel(self.model)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setColumnHidden(0, True)
        self.table.refresh_style()
    
    def export_to_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            QMessageBox.warning(self, "تنبيه", "مكتبة openpyxl غير مثبتة")
            return
        filename, _ = QFileDialog.getSaveFileName(self, "حفظ سجل التدقيق", "audit_log.xlsx", "Excel (*.xlsx)")
        if not filename:
            return
        all_logs = self.repo.export_all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "سجل التدقيق"
        headers = ['id', 'user_id', 'username', 'action', 'table_name', 'record_id', 'details', 'ip_address', 'timestamp']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row, log in enumerate(all_logs, 2):
            for col, key in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=str(log.get(key, '')))
        wb.save(filename)
        QMessageBox.information(self, "نجاح", f"تم التصدير إلى {filename}")
    
    def delete_old_logs(self):
        reply = QMessageBox.question(self, "تأكيد الحذف", "هل أنت متأكد من حذف جميع السجلات الأقدم من 90 يوماً؟ لا يمكن التراجع.", QMessageBox.Yes|QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.repo.delete_old_logs(90)
            QMessageBox.information(self, "نجاح", "تم حذف السجلات القديمة")
            self.refresh()
    
    def show_stats(self):
        stats = self.repo.get_stats()
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QTabWidget
        dialog = QDialog(self)
        dialog.setWindowTitle("إحصائيات سجل التدقيق")
        dialog.setLayoutDirection(Qt.RightToLeft)  # RTL
        dialog.resize(600, 400)
        layout = QVBoxLayout(dialog)
        tabs = QTabWidget()
        tabs.setLayoutDirection(Qt.RightToLeft)  # RTL
        
        user_table = QTableWidget()
        user_table.setLayoutDirection(Qt.RightToLeft)
        user_table.setColumnCount(2)
        user_table.setHorizontalHeaderLabels(["المستخدم", "عدد العمليات"])
        user_table.setRowCount(len(stats['by_user']))
        for i, row in enumerate(stats['by_user']):
            user_table.setItem(i, 0, QTableWidgetItem(row['username']))
            user_table.setItem(i, 1, QTableWidgetItem(str(row['count'])))
        tabs.addTab(user_table, "حسب المستخدم")
        
        action_table = QTableWidget()
        action_table.setLayoutDirection(Qt.RightToLeft)
        action_table.setColumnCount(2)
        action_table.setHorizontalHeaderLabels(["العملية", "العدد"])
        action_table.setRowCount(len(stats['by_action']))
        for i, row in enumerate(stats['by_action']):
            action_table.setItem(i, 0, QTableWidgetItem(row['action']))
            action_table.setItem(i, 1, QTableWidgetItem(str(row['count'])))
        tabs.addTab(action_table, "حسب العملية")
        
        table_table = QTableWidget()
        table_table.setLayoutDirection(Qt.RightToLeft)
        table_table.setColumnCount(2)
        table_table.setHorizontalHeaderLabels(["الجدول", "العدد"])
        table_table.setRowCount(len(stats['by_table']))
        for i, row in enumerate(stats['by_table']):
            table_table.setItem(i, 0, QTableWidgetItem(row['table_name']))
            table_table.setItem(i, 1, QTableWidgetItem(str(row['count'])))
        tabs.addTab(table_table, "حسب الجدول")
        
        daily_table = QTableWidget()
        daily_table.setLayoutDirection(Qt.RightToLeft)
        daily_table.setColumnCount(2)
        daily_table.setHorizontalHeaderLabels(["التاريخ", "العدد"])
        daily_table.setRowCount(len(stats['daily']))
        for i, row in enumerate(stats['daily']):
            daily_table.setItem(i, 0, QTableWidgetItem(row['day']))
            daily_table.setItem(i, 1, QTableWidgetItem(str(row['count'])))
        tabs.addTab(daily_table, "حسب اليوم (آخر 30 يوماً)")
        
        layout.addWidget(tabs)
        close_btn = QPushButton("إغلاق")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        dialog.exec()
    
    def print_audit_log(self):
        from printing.print_manager import PrintManager
        model = self.table.model()
        if not model or model.rowCount() == 0:
            QMessageBox.warning(self, "تنبيه", "لا توجد بيانات للطباعة")
            return
        headers = [model.headerData(i, Qt.Horizontal, Qt.DisplayRole) for i in range(model.columnCount())]
        data = []
        for row in range(model.rowCount()):
            row_data = []
            for col in range(model.columnCount()):
                idx = model.index(row, col)
                value = model.data(idx, Qt.DisplayRole)
                row_data.append(str(value) if value else '')
            data.append(row_data)
        PrintManager.print_report("سجل التدقيق", headers, data, self)
