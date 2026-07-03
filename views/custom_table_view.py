from PyQt5.QtWidgets import QTableView, QMenu, QAction, QFileDialog, QMessageBox, QApplication, QHeaderView, QStyledItemDelegate
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QKeySequence, QTextDocument, QTextOption, QFont
from PyQt5.QtPrintSupport import QPrinter, QPrintPreviewDialog
from theme_manager import ThemeManager
from services.print_service import print_service

class CenterAlignDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        option.displayAlignment = Qt.AlignCenter
        super().paint(painter, option, index)

class CustomTableView(QTableView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setLayoutDirection(Qt.RightToLeft)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QTableView.SelectRows)
        self.setSelectionMode(QTableView.ExtendedSelection)
        self.setSortingEnabled(True)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_menu)

        self.setItemDelegate(CenterAlignDelegate(self))

        self.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.verticalHeader().setDefaultAlignment(Qt.AlignCenter)

        self.copy_action = QAction(self)
        self.copy_action.setShortcut(QKeySequence.Copy)
        self.copy_action.triggered.connect(self.copy_selection)
        self.addAction(self.copy_action)

        self.refresh_style()

    def copy_selection(self):
        selection = self.selectionModel().selectedIndexes()
        if not selection:
            return
        rows = sorted(set(i.row() for i in selection))
        cols = sorted(set(i.column() for i in selection))
        text = ""
        for r in rows:
            row_data = []
            for c in cols:
                idx = self.model().index(r, c)
                data = self.model().data(idx, Qt.DisplayRole)
                row_data.append(str(data))
            text += "\t".join(row_data) + "\n"
        QApplication.clipboard().setText(text)

    def _show_menu(self, pos):
        menu = QMenu()
        export_excel = QAction("📊 تصدير إلى Excel", self)
        export_excel.triggered.connect(self.export_to_excel)
        menu.addAction(export_excel)
        export_pdf = QAction("📄 طباعة", self)
        export_pdf.triggered.connect(self.print_table)
        menu.addAction(export_pdf)
        menu.addSeparator()
        copy_act = QAction("📋 نسخ", self)
        copy_act.triggered.connect(self.copy_selection)
        menu.addAction(copy_act)
        menu.exec(self.viewport().mapToGlobal(pos))

    def export_to_excel(self):
        model = self.model()
        if not model:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            QMessageBox.warning(self, "تنبيه", "مكتبة openpyxl غير مثبتة")
            return
        filename, _ = QFileDialog.getSaveFileName(self, "حفظ التقرير", "report.xlsx", "Excel (*.xlsx)")
        if not filename:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير"
        for col in range(model.columnCount()):
            header = model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
            cell = ws.cell(row=1, column=col+1, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in range(model.rowCount()):
            for col in range(model.columnCount()):
                idx = model.index(row, col)
                value = model.data(idx, Qt.DisplayRole)
                ws.cell(row=row+2, column=col+1, value=value)
        wb.save(filename)
        QMessageBox.information(self, "نجاح", f"تم التصدير إلى {filename}")

    def print_table(self):
        model = self.model()
        if not model:
            return
        headers = []
        for col in range(model.columnCount()):
            h = model.headerData(col, Qt.Horizontal, Qt.DisplayRole)
            headers.append(str(h) if h else f"عمود{col+1}")
        data = []
        for row in range(model.rowCount()):
            row_data = []
            for col in range(model.columnCount()):
                idx = model.index(row, col)
                val = model.data(idx, Qt.DisplayRole)
                row_data.append(str(val) if val is not None else '')
            data.append(row_data)

        html = print_service.build_table_report('طباعة الجدول', headers, data)
        doc = QTextDocument()
        doc.setHtml(html)
        doc.setDefaultFont(QFont("Tajawal", 10))
        opt = QTextOption()
        opt.setAlignment(Qt.AlignRight)
        opt.setTextDirection(Qt.RightToLeft)
        doc.setDefaultTextOption(opt)
        printer = QPrinter(QPrinter.HighResolution)
        preview = QPrintPreviewDialog(printer, self)
        preview.setLayoutDirection(Qt.RightToLeft)
        preview.paintRequested.connect(lambda p: doc.print_(p))
        preview.exec()

    def refresh_style(self):
        self.setStyleSheet(ThemeManager.get_stylesheet())
        self.viewport().update()

    def showEvent(self, event):
        self.refresh_style()
        super().showEvent(event)
