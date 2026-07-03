# -*- coding: utf-8 -*-
"""Report export helpers for CSV, XLSX and HTML.

The service works with already-calculated rows. It must not recalculate financial
values, so exported files match the inline preview exactly.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Iterable, List, Sequence, Optional


class ExportService:
    INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1F]')

    def safe_filename(self, title: str, extension: str) -> str:
        base = self.INVALID_FILENAME_CHARS.sub('_', title or 'report').strip(' ._') or 'report'
        base = re.sub(r'\s+', '_', base)
        ext = extension if extension.startswith('.') else f'.{extension}'
        return f'{base}{ext}'

    def normalize_rows(self, rows: Iterable[Sequence[Any]]) -> List[List[str]]:
        return [["" if cell is None else str(cell) for cell in row] for row in rows]

    def write_html(self, path: str | Path, html: str) -> Path:
        output = Path(path)
        if output.suffix.lower() != '.html':
            output = output.with_suffix('.html')
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(html or '', encoding='utf-8')
        return output

    def write_csv(self, path: str | Path, headers: Sequence[Any], rows: Iterable[Sequence[Any]]) -> Path:
        output = Path(path)
        if output.suffix.lower() != '.csv':
            output = output.with_suffix('.csv')
        output.parent.mkdir(parents=True, exist_ok=True)
        with output.open('w', encoding='utf-8-sig', newline='') as fh:
            writer = csv.writer(fh)
            writer.writerow(["" if h is None else str(h) for h in headers])
            writer.writerows(self.normalize_rows(rows))
        return output

    def write_xlsx(
        self,
        path: str | Path,
        title: str,
        headers: Sequence[Any],
        rows: Iterable[Sequence[Any]],
        subtitle: Optional[str] = None,
    ) -> Path:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - depends on runtime packaging
            raise RuntimeError('مكتبة openpyxl غير مثبتة. ثبّت المتطلبات أو صدّر CSV بدل Excel.') from exc

        output = Path(path)
        if output.suffix.lower() != '.xlsx':
            output = output.with_suffix('.xlsx')
        output.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'تقرير'
        ws.sheet_view.rightToLeft = True

        row_idx = 1
        ws.cell(row=row_idx, column=1, value=title or 'تقرير')
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=15)
        row_idx += 1
        if subtitle:
            ws.cell(row=row_idx, column=1, value=subtitle)
            ws.cell(row=row_idx, column=1).font = Font(color='666666')
            row_idx += 1
        row_idx += 1

        header_fill = PatternFill(fill_type='solid', fgColor='E5E7EB')
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="" if header is None else str(header))
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        data_start = row_idx + 1
        for r_idx, row in enumerate(self.normalize_rows(rows), data_start):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)

        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 10
            for cell in ws[letter]:
                max_len = max(max_len, len(str(cell.value or '')))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

        ws.freeze_panes = ws.cell(row=data_start, column=1)
        wb.save(output)
        return output


export_service = ExportService()
